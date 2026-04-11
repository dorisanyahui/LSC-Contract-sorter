import re
import openpyxl
from pathlib import Path

_CITIES = [
    "上海", "北京", "广州", "深圳", "天津", "重庆",
    "沈阳", "青岛", "无锡", "苏州", "成都", "武汉",
    "杭州", "南京", "西安", "大连", "宁波", "厦门",
    "常州", "潍坊", "长沙", "郑州", "济南", "合肥",
]

def strip_city(name: str) -> str:
    """去除公司名中的城市信息，方便跨格式匹配。
    例：青岛派克汉尼汾流体连接件有限公司 → 派克汉尼汾流体连接件有限公司
        派克汉尼汾流体连接件(青岛)有限公司 → 派克汉尼汾流体连接件有限公司
    """
    for city in _CITIES:
        name = re.sub(r"\(" + city + r"\)", "", name)
        name = re.sub(r"（" + city + r"）", "", name)
        if name.startswith(city):
            name = name[len(city):]
    return name.strip()


def normalize_name(name: str) -> str:
    if not name:
        return ""
    name = str(name).strip()
    name = name.replace("（", "(").replace("）", ")")
    name = name.replace("【", "[").replace("】", "]")
    name = name.replace("〈", "(").replace("〉", ")")
    name = name.replace("《", "").replace("》", "")
    name = name.replace("\u3000", " ")
    name = "".join(name.split())
    return name.lower()


def load_group_mapping_from_excel(excel_path: Path, sheet_name: str = "mapping_clean") -> dict:
    """
    从 group_company_mapping_clean.xlsx 读取集团映射，生成：
    {
      "集团A": {
        "companies": [...],
        "aliases": [...]
      },
      ...
    }
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"未找到映射表 Excel: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel 中未找到工作表: {sheet_name}")

    ws = wb[sheet_name]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_index = {str(v).strip(): i + 1 for i, v in enumerate(headers) if v is not None}

    required_cols = ["名称", "所属集团名称", "简称"]
    for col in required_cols:
        if col not in header_index:
            raise ValueError(f"映射表缺少必要列: {col}")

    mapping = {}

    for row in range(2, ws.max_row + 1):
        company = ws.cell(row, header_index["名称"]).value
        group_name = ws.cell(row, header_index["所属集团名称"]).value
        short_name = ws.cell(row, header_index["简称"]).value

        company = str(company).strip() if company else ""
        group_name = str(group_name).strip() if group_name else ""
        short_name = str(short_name).strip() if short_name else ""

        if not company or not group_name:
            continue

        if group_name not in mapping:
            mapping[group_name] = {
                "companies": [],
                "aliases": []
            }

        if company not in mapping[group_name]["companies"]:
            mapping[group_name]["companies"].append(company)

        if short_name and short_name not in mapping[group_name]["aliases"]:
            mapping[group_name]["aliases"].append(short_name)

    return mapping


def build_company_aliases_from_group_mapping(group_mapping: dict) -> dict:
    """
    转成 main.py 里继续可用的结构：
    {
      "公司A": ["集团简称1", "集团简称2"],
      "公司B": [...]
    }
    """
    company_aliases = {}

    for group_name, info in group_mapping.items():
        if not isinstance(info, dict):
            continue

        companies = info.get("companies", [])
        aliases = info.get("aliases", [])

        if not isinstance(companies, list):
            companies = []
        if not isinstance(aliases, list):
            aliases = []

        clean_aliases = []
        for a in aliases:
            if a is None:
                continue
            a = str(a).strip()
            if a:
                clean_aliases.append(a)

        # 也把集团名本身加入 alias，方便匹配
        if group_name and group_name not in clean_aliases:
            clean_aliases.append(group_name)

        for company in companies:
            if company is None:
                continue
            company = str(company).strip()
            if not company:
                continue

            if company not in company_aliases:
                company_aliases[company] = []

            for a in clean_aliases:
                if a not in company_aliases[company] and a != company:
                    company_aliases[company].append(a)

    return company_aliases


def map_company_to_group(company_name: str, group_mapping: dict) -> str:
    """
    根据识别出的公司名，映射到所属集团。
    匹配顺序：
    1. 公司全称标准化精确匹配
    2. 公司全称包含匹配
    3. 集团 aliases 辅助匹配
    """
    if not company_name:
        return ""

    company_n = normalize_name(company_name)
    company_n_no_city = normalize_name(strip_city(company_name))

    exact_hits = []
    fuzzy_hits = []
    alias_hits = []

    for group_name, info in group_mapping.items():
        if not isinstance(info, dict):
            continue

        companies = info.get("companies", [])
        aliases = info.get("aliases", [])

        if not isinstance(companies, list):
            companies = []
        if not isinstance(aliases, list):
            aliases = []

        for comp in companies:
            comp_n = normalize_name(comp)
            comp_n_no_city = normalize_name(strip_city(comp))
            if not comp_n:
                continue

            if company_n == comp_n:
                exact_hits.append(group_name)
            elif company_n_no_city and comp_n_no_city and company_n_no_city == comp_n_no_city:
                exact_hits.append(group_name)  # 去城市后完全匹配
            elif len(company_n) >= 4 and len(comp_n) >= 4 and (
                company_n in comp_n or comp_n in company_n
            ):
                fuzzy_hits.append(group_name)

        for alias in aliases:
            alias_n = normalize_name(alias)
            if not alias_n or len(alias_n) < 4:
                continue

            if company_n == alias_n:
                exact_hits.append(group_name)
            elif len(company_n) >= 4 and (company_n in alias_n or alias_n in company_n):
                alias_hits.append(group_name)

        group_n = normalize_name(group_name)
        if group_n and len(group_n) >= 4 and (
            company_n == group_n or company_n in group_n or group_n in company_n
        ):
            alias_hits.append(group_name)

    if exact_hits:
        return exact_hits[0]
    if fuzzy_hits:
        return fuzzy_hits[0]
    if alias_hits:
        return alias_hits[0]

    return ""