from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models.enums import DocType
from src.models.schema import DocumentResult
from src.utils.file_utils import ensure_dir


# Doc types that are expected to carry monetary amounts
AMOUNT_DOC_TYPES = {DocType.CONTRACT, DocType.SRF, DocType.PURCHASE_ORDER, DocType.QUOTE}


def _has_amount(result: DocumentResult) -> bool:
    return any([
        result.annual_maintenance_fee,
        result.tax_included_amount,
        result.tax_excluded_amount,
        result.contract_total_amount,
    ])


def _filter_amount_results(results: list[DocumentResult]) -> list[DocumentResult]:
    """Keep only docs of relevant types that have at least one amount field."""
    return [r for r in results if r.doc_type in AMOUNT_DOC_TYPES and _has_amount(r)]


COLUMN_NAMES = [
    "年份",
    "集团名称",
    "公司名称",
    "对手方",
    "文档类型",
    "文档子类型",
    "是否主文档",
    "签署日期",
    "生效日期",
    "服务开始",
    "服务结束",
    "年度维护金额",
    "税率",
    "含税金额",
    "不含税金额",
    "合同订单总金额",
    "币种",
    "PO编号",
    "报价编号",
    "文件名",
    "文件路径",
    "MD5",
    "摘要",
    "置信度",
    "flags",
    "错误信息",
]


class ExcelExporter:
    """Exports DocumentResult lists to Excel files."""

    def _result_to_row(self, result: DocumentResult) -> dict:
        """Convert a DocumentResult to a flat dict for Excel export."""
        return {
            "年份": result.report_year,
            "集团名称": result.detected_group,
            "公司名称": result.detected_company,
            "对手方": result.detected_counterparty,
            "文档类型": result.doc_type.value if result.doc_type else "",
            "文档子类型": result.doc_subtype,
            "是否主文档": "是" if result.is_primary_doc else "否",
            "签署日期": result.sign_date,
            "生效日期": result.effective_date,
            "服务开始": result.service_period_start,
            "服务结束": result.service_period_end,
            "年度维护金额": result.annual_maintenance_fee,
            "税率": result.tax_rate,
            "含税金额": result.tax_included_amount,
            "不含税金额": result.tax_excluded_amount,
            "合同订单总金额": result.contract_total_amount,
            "币种": result.currency.value if result.currency else "CNY",
            "PO编号": result.po_number,
            "报价编号": result.quote_number,
            "文件名": result.file_name,
            "文件路径": result.file_path,
            "MD5": result.file_md5,
            "摘要": result.summary,
            "置信度": round(result.confidence_overall, 3),
            "flags": "|".join(result.flags) if result.flags else "",
            "错误信息": result.error or "",
        }

    def _write_dataframe_to_excel(self, df: pd.DataFrame, output_path: Path, sheet_name: str = "数据") -> None:
        """Write a DataFrame to a formatted Excel file."""
        ensure_dir(output_path.parent)
        df.to_excel(output_path, index=False, sheet_name=sheet_name, engine="openpyxl")

        # Apply formatting
        try:
            wb = load_workbook(output_path)
            ws = wb.active

            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-width columns
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    try:
                        cell_len = len(str(cell.value)) if cell.value is not None else 0
                        if cell_len > max_len:
                            max_len = cell_len
                    except Exception:
                        pass
                adjusted_width = min(max_len + 4, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(output_path)
        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting to {output_path}: {e}")

    def export_all(self, results: list[DocumentResult], output_dir: Path) -> None:
        """Export all results to a single all_documents.xlsx."""
        ensure_dir(output_dir)
        rows = [self._result_to_row(r) for r in results]
        df = pd.DataFrame(rows, columns=COLUMN_NAMES)
        output_path = output_dir / "all_documents.xlsx"
        self._write_dataframe_to_excel(df, output_path, sheet_name="全部文件")
        logger.info(f"Exported {len(results)} documents to {output_path}")

        # Export filtered summary: only amount-bearing doc types with at least one amount
        filtered = _filter_amount_results(results)
        if filtered:
            filtered_rows = [self._result_to_row(r) for r in filtered]
            filtered_df = pd.DataFrame(filtered_rows, columns=COLUMN_NAMES)
            filtered_path = output_dir / "有金额汇总.xlsx"
            self._write_dataframe_to_excel(filtered_df, filtered_path, sheet_name="有金额文件")
            logger.info(f"Exported {len(filtered)} amount docs to {filtered_path}")

        # Also archive files to output/{group}/{year}/
        self._archive_files(results, output_dir)

    def export_by_group(self, results: list[DocumentResult], output_dir: Path) -> None:
        """Export results grouped by group name to separate Excel files."""
        groups: dict[str, list[DocumentResult]] = {}
        for result in results:
            group = result.detected_group or "未分组"
            groups.setdefault(group, []).append(result)

        for group_name, group_results in groups.items():
            group_dir = output_dir / group_name
            ensure_dir(group_dir)

            rows = [self._result_to_row(r) for r in group_results]
            df = pd.DataFrame(rows, columns=COLUMN_NAMES)

            safe_name = group_name.replace("/", "_").replace("\\", "_")
            output_path = group_dir / f"{safe_name}_汇总表.xlsx"
            self._write_dataframe_to_excel(df, output_path, sheet_name=group_name[:30])
            logger.info(f"Exported {len(group_results)} rows for group '{group_name}' to {output_path}")

            # Per-group filtered summary
            filtered = _filter_amount_results(group_results)
            if filtered:
                filtered_rows = [self._result_to_row(r) for r in filtered]
                filtered_df = pd.DataFrame(filtered_rows, columns=COLUMN_NAMES)
                filtered_path = group_dir / f"{safe_name}_有金额汇总.xlsx"
                self._write_dataframe_to_excel(filtered_df, filtered_path, sheet_name=group_name[:30])
                logger.info(f"Exported {len(filtered)} amount docs for group '{group_name}'")

    def export_packets(self, results: list[DocumentResult], output_dir: Path) -> None:
        """Export packet-level data to packets.xlsx."""
        ensure_dir(output_dir)
        rows = []
        for result in results:
            for packet in result.packets:
                rows.append({
                    "文件名": result.file_name,
                    "集团": result.detected_group,
                    "packet_id": packet.packet_id,
                    "start_page": packet.start_page,
                    "end_page": packet.end_page,
                    "packet_type": packet.packet_type,
                    "doc_type": packet.doc_type.value,
                    "title_hint": packet.title_hint[:100],
                })

        if rows:
            df = pd.DataFrame(rows)
            output_path = output_dir / "packets.xlsx"
            self._write_dataframe_to_excel(df, output_path, sheet_name="Packets")
            logger.info(f"Exported {len(rows)} packets to {output_path}")

    def _archive_files(self, results: list[DocumentResult], output_dir: Path) -> None:
        """Archive source files to output/{group}/{year}/ directories."""
        for result in results:
            if not result.file_path or result.error:
                continue

            src = Path(result.file_path)
            if not src.exists():
                continue

            group = result.detected_group or "未分组"
            year = str(result.report_year) if result.report_year else "未知年份"

            dst_dir = output_dir / group / year
            ensure_dir(dst_dir)

            dst = dst_dir / src.name
            if not dst.exists():
                try:
                    shutil.copy2(src, dst)
                except Exception as e:
                    logger.warning(f"Failed to archive {src} -> {dst}: {e}")
